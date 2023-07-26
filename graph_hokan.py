import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from scipy import interpolate
import scipy as sp


file_path = 'C:/Users/mpg/Desktop/python_excel/test.xlsx'
interpolate_num = 300
# グラフに使うデータを取得します
x_values = []
y_values = []


workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

def get_last_row_with_data(sheet):
    max_row = sheet.max_row
    for row in range(max_row, 0, -1):
        cell_value = sheet.cell(row=row, column=3).value
        if cell_value is not None and cell_value != "":
            print("最終行の値は、"+str(row))
            print("最終行の数は、"+str(cell_value))
            return row
    return None

def find_nearest_larger_value(dateNumber):
    second_exponentiation=[256,512,1024,2048,4096,8192,16384,32768,65536,131072,262144,524288]
    nearest_larger_num = float('inf')  # 正の無限大で初期化
    for num in second_exponentiation:
        if dateNumber < num < nearest_larger_num:
            nearest_larger_num = num
    return nearest_larger_num

dateNum=get_last_row_with_data(sheet)
interpolate_num=find_nearest_larger_value(dateNum)
print("1番近くて大きな2の累乗は、↓")
print(interpolate_num)

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=3, max_col=4):#列3と列4を指定
    x_values.append(row[0].value)  # 1列目の値をx軸に
    y_values.append(row[1].value)  # 2列目の値をy軸に

myfunc = sp.interpolate.interp1d(x_values, y_values)
# x軸のデータを300に補間後
x_new = np.linspace(min(x_values), max(x_values), interpolate_num)

#print(x_new)
#print(myfunc(x_new))

plt.plot(x_values, y_values)
plt.xlabel('時間 [s]',fontsize=10, fontname='Meiryo')
plt.ylabel('グレースケール値 ',fontsize=10, fontname='Meiryo')
plt.title('グレースケールのグラフ',fontsize=14, fontname='Meiryo')
plt.grid(True)
plt.show()

#plt.scatter(x_values, y_values, c = 'b', label = 'original')
plt.scatter(x_new, myfunc(x_new), facecolor = 'None', edgecolors = 'r', label = 'new')
plt.xlabel('時間 [s]',fontsize=10, fontname='Meiryo')
plt.ylabel('グレースケール値の補間後 ',fontsize=10, fontname='Meiryo')
plt.title('グレースケールを'+str(interpolate_num)+'個に補間',fontsize=14, fontname='Meiryo')
plt.legend(bbox_to_anchor = (1, 1))
plt.grid()
plt.show()

# 変更を保存します
workbook.save(file_path)

# メモリを解放します
workbook.close()


