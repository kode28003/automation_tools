import openpyxl
import matplotlib.pyplot as plt


file_path = 'C:/Users/mpg/Desktop/python_excel/test.xlsx'

def get_last_value_in_column(sheet, column):
    last_row = sheet.max_row
    while last_row > 0:
        cell_value = sheet[f"{column}{last_row}"].value
        if cell_value is not None and cell_value != "":
            return cell_value
        last_row -= 1
    return None

def findEmptyColumn():
    for row in sheet.iter_rows(min_row=2, max_row=2, min_col=1, max_col=30):
        for cell in row:
            if cell.value is None:
                return cell.column        

def correctTimeFromMatlabTime(experimentDuration,last_value_A,writeCol):
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        #sheet.iter_rows(最小の縦番号、最大の縦番号、最小の列、最大の列) col=1=A -- col=2=B
        for cell in row:
            if cell.value is not None:  # もしセルに値がある場合のみ処理を行います
                new_value = (cell.value /last_value_A)*experimentDuration
                sheet.cell(row=cell.row, column=writeCol, value=new_value)
    
def copyColumn(targetColumn):
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=2, max_col=2):
        for cell in row:
            sheet.cell(row=cell.row, column=targetColumn, value=cell.value)

workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# A列の1番下の値を取得します
last_value_A = get_last_value_in_column(sheet, 'A')
print(f"A列の1番下の値: {last_value_A}")

nonCellNum1=findEmptyColumn()
print("空いている行1は、"+str(nonCellNum1))

experimentSec=input("実験時間を秒数で入力してください")
correctTimeFromMatlabTime(int(experimentSec),last_value_A,nonCellNum1)
nonCellNum2=findEmptyColumn()
print("空いている行2は、"+str(nonCellNum1))
copyColumn(nonCellNum2)

# グラフに使うデータを取得します
# x_values = []
# y_values = []
# for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=3, max_col=4):
#     x_values.append(row[0].value)  # 1列目の値をx軸に
#     y_values.append(row[1].value)  # 2列目の値をy軸に

# # グラフを作成して表示します
# plt.plot(x_values, y_values)
# plt.xlabel('時間 [s]',fontsize=15, fontname='Meiryo')
# plt.ylabel('グレースケール値 ',fontsize=15, fontname='Meiryo')
# #plt.title('グレースケールー時間',fontsize=15, fontname='Meiryo')
# plt.grid(True)
# plt.show()

# 変更を保存します
workbook.save(file_path)

# メモリを解放します
workbook.close()



