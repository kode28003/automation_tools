import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from scipy import interpolate
import scipy as sp
from scipy import signal
from scipy import fftpack
from scipy.signal import butter, filtfilt


def get_last_row_with_data(sheet):
    max_row = sheet.max_row
    for row in range(max_row, 0, -1):
        cell_value = sheet.cell(row=row, column=3).value
        if cell_value is not None and cell_value != "":
            print("最終行の値は、"+str(row))
            print("最終行の数は、"+str(cell_value))
            return row
    return None

def find_nearest_larger_value(dateNumber):
    two_exponentiation=[256,512,1024,2048,4096,8192,16384,32768,65536,131072,262144,524288]
    nearest_larger_num = float('inf')  # 正の無限大で初期化
    for num in two_exponentiation:
        if dateNumber < num < nearest_larger_num:
            nearest_larger_num = num
    return nearest_larger_num

def butter_bandpass(lowcut, highcut, fs, order=5):
    nyquist = 0.5 * fs
    low = lowcut / nyquist
    high = highcut / nyquist
    b, a = butter(order, [low, high], btype='band')
    return b, a

def bandpass_filter(data, lowcut, highcut, fs, order=5):
    b, a = butter_bandpass(lowcut, highcut, fs, order=order)
    filtered_data = filtfilt(b, a, data)
    return filtered_data

x_values = []
y_values = []

file_path = 'C:/Users/mpg/Desktop/python_excel/test.xlsx'
interpolate_num = 256 #補間の初期値

workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

dateNum=get_last_row_with_data(sheet)
interpolate_num=find_nearest_larger_value(dateNum)
print("1番近いかつ大きな2の累乗は、↓")
print(interpolate_num)

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=3, max_col=4):#列3と列4を指定
    x_values.append(row[0].value)  # 3列目の値をx軸に
    y_values.append(row[1].value)  # 4列目の値をy軸に

myfunc = sp.interpolate.interp1d(x_values, y_values)
x_new = np.linspace(min(x_values), max(x_values), interpolate_num)#まず、xを2の累乗に補間
print(x_new)

fs = 1 / (x_new[1] - x_new[0])  # サンプリング周波数 (Hz)
lowcut = 0.75  # 0.75Hz以下の周波数をカット
highcut = 2.5  # 2.5Hz以上の周波数をカット
filtered_data = bandpass_filter(myfunc(x_new), lowcut, highcut, fs, order=6)

plt.figure(figsize=(10, 6))

plt.plot(x_new, myfunc(x_new), 'b-', label='Original Data')
plt.xlabel('Time [s]')
plt.ylabel('Amplitude')
plt.legend()
plt.grid(True)
plt.show()

plt.plot(x_new, filtered_data, 'r-', linewidth=2, label='Filtered Data')
plt.xlabel('Time [s]')
plt.ylabel('Amplitude')
plt.legend()
plt.grid(True)
plt.xlim(40, 60)
plt.ylim(-1, 1)
plt.tight_layout()
plt.show()

x_peaks =  [x_new[i] for i in range(1, len(x_new)-1) if filtered_data[i] > 0 and filtered_data[i] > filtered_data[i-1] and filtered_data[i] > filtered_data[i+1]]
y_peaks = [filtered_data[i] for i in range(1, len(filtered_data)-1) if filtered_data[i] > 0 and filtered_data[i] > filtered_data[i-1] and filtered_data[i] > filtered_data[i+1]]

# プロット
plt.plot(x_new, filtered_data, 'r-', linewidth=2, label='Filtered Data')
plt.plot(x_peaks, y_peaks, marker='o', linestyle='', label='peak', color='blue')
plt.xlabel('Time [s]')
plt.ylabel('Amplitude')
plt.legend()
plt.grid(True)
plt.xlim(40, 60)
plt.ylim(-1, 1)
plt.show()